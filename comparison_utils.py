"""
Utilities for comparing multiple optimization runs.

Provides functions for:
- Discovering result directories from path patterns
- Loading and validating CSV files
- Generating case names from directory paths
- Creating Excel comparison workbooks
"""

import glob
from pathlib import Path
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


def discover_result_directories(path_patterns):
    """
    Discover valid result directories from path patterns.

    Expands glob patterns and validates that each directory contains
    the required optimization_summary.csv file.

    Parameters
    ----------
    path_patterns : list of str
        List of directory paths or glob patterns
        Examples: ['results/baseline/', 'results/test_*/']

    Returns
    -------
    list of Path
        Sorted list of directories containing optimization_summary.csv

    Raises
    ------
    ValueError
        If no valid directories found

    Examples
    --------
    >>> dirs = discover_result_directories(['results/test_*_000/'])
    >>> dirs = discover_result_directories(['results/baseline/', 'results/high_eta/'])
    """
    directories = []

    for pattern in path_patterns:
        matches = glob.glob(pattern)
        for match in matches:
            path = Path(match)
            if path.is_dir() and (path / 'optimization_summary.csv').exists():
                directories.append(path)

    if not directories:
        raise ValueError(f"No valid result directories found for patterns: {path_patterns}")

    return sorted(set(directories))


def generate_case_name(directory_path):
    """
    Generate readable case name from directory path.

    Uses the directory name as the case name for display in plots and tables.

    Parameters
    ----------
    directory_path : Path
        Directory path

    Returns
    -------
    str
        Case name for display

    Examples
    --------
    >>> generate_case_name(Path('results/baseline'))
    'baseline'
    >>> generate_case_name(Path('results/test_010'))
    'test_010'
    """
    return directory_path.name


def load_optimization_summaries(directories):
    """
    Load optimization_summary.csv from multiple directories.

    Parameters
    ----------
    directories : list of Path
        Result directories

    Returns
    -------
    dict
        {case_name: pd.DataFrame, ...}
        Each DataFrame contains optimization summary with columns:
        iteration, n_evaluations, objective, termination_status, ...
        (elapsed_time included if available)

    Examples
    --------
    >>> dirs = [Path('results/baseline'), Path('results/test_010')]
    >>> data = load_optimization_summaries(dirs)
    >>> data.keys()
    dict_keys(['baseline', 'test_010'])
    """
    data = {}
    for directory in directories:
        case_name = generate_case_name(directory)
        csv_path = directory / 'optimization_summary.csv'

        # Parse custom CSV format - find iteration history section
        with open(csv_path, 'r') as f:
            lines = f.readlines()

        # Find the "Iterative Refinement - Iteration History" section
        start_idx = None
        for i, line in enumerate(lines):
            if 'Iterative Refinement - Iteration History' in line:
                start_idx = i + 1  # Skip section header, header is next line
                break

        if start_idx is None:
            raise ValueError(f"Could not find iteration history in {csv_path}")

        # Read header and data rows
        header_line = lines[start_idx].strip()
        data_start = start_idx + 1

        # Read rows until blank line or next section
        data_rows = []
        for i in range(data_start, len(lines)):
            line = lines[i].strip()
            if not line or line.startswith('Iterative'):
                break
            data_rows.append(line)

        # Parse into DataFrame
        from io import StringIO
        csv_content = header_line + '\n' + '\n'.join(data_rows)
        df = pd.read_csv(StringIO(csv_content))

        # Rename columns to match expected format
        column_mapping = {
            'Iteration': 'iteration',
            'Objective': 'objective',
            'Evaluations': 'n_evaluations',
            'Status': 'termination_status'
        }
        df = df.rename(columns=column_mapping)

        data[case_name] = df

    return data


def load_results_csvs(directories):
    """
    Load results.csv from multiple directories.

    Only includes cases where results.csv exists. Prints warning for
    directories missing results.csv.

    Parameters
    ----------
    directories : list of Path
        Result directories

    Returns
    -------
    dict
        {case_name: pd.DataFrame, ...}
        Only includes cases where results.csv exists.
        Each DataFrame contains model results with columns:
        t, y, K, T_atm, E_cum, etc.

    Examples
    --------
    >>> dirs = [Path('results/baseline'), Path('results/test_010')]
    >>> data = load_results_csvs(dirs)
    """
    data = {}
    for directory in directories:
        case_name = generate_case_name(directory)
        csv_path = directory / 'results.csv'
        if csv_path.exists():
            data[case_name] = pd.read_csv(csv_path)
        else:
            print(f"Warning: results.csv not found in {directory}, skipping results comparison for this case")

    return data


def create_directories_sheet(wb, directories):
    """
    Create sheet listing all directories included in comparison.

    Parameters
    ----------
    wb : openpyxl.Workbook
        Workbook to add sheet to
    directories : list of Path
        List of result directories
    """
    ws = wb.create_sheet('Directories')

    # Header row
    ws['A1'] = 'Case Name'
    ws['A1'].font = Font(bold=True)
    ws['A1'].fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')

    ws['B1'] = 'Directory Path'
    ws['B1'].font = Font(bold=True)
    ws['B1'].fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')

    # Data rows
    for row_idx, directory in enumerate(directories, start=2):
        case_name = generate_case_name(directory)
        ws.cell(row_idx, 1, case_name)
        ws.cell(row_idx, 2, str(directory))


def create_comparison_xlsx(optimization_data, directories, output_path):
    """
    Create Excel workbook comparing optimization summaries across cases.

    Creates multi-sheet workbook with one sheet per metric (Objective,
    Evaluations, Elapsed Time, Termination Status). Cases are columns,
    iterations are rows.

    Parameters
    ----------
    optimization_data : dict
        {case_name: optimization_summary_df, ...}
    directories : list of Path
        List of result directories included in comparison
    output_path : Path or str
        Output Excel file path

    Notes
    -----
    Workbook structure:
    - Sheet 1: "Directories" - list of all compared directories
    - Sheet 2: "Objective" - objective values by iteration
    - Sheet 3: "Evaluations" - function evaluation counts
    - Sheet 4: "Elapsed Time (s)" - computation time (if available)
    - Sheet 5: "Termination Status" - optimization termination reasons

    Each metric sheet has:
    - Column A: Iteration number
    - Columns B+: One column per case
    - Header row with case names

    Examples
    --------
    >>> create_comparison_xlsx(opt_data, dirs, 'comparison_summary.xlsx')
    Comparison Excel workbook saved to: comparison_summary.xlsx
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Sheet 1: List of directories
    create_directories_sheet(wb, directories)

    # Define comparison sheets with column name and number format
    comparison_specs = [
        ('Objective', 'objective', '{:.6e}'),
        ('Evaluations', 'n_evaluations', '{:.0f}'),
        ('Termination Status', 'termination_status', '{}')
    ]

    # Add elapsed_time sheet only if data contains it
    if optimization_data and any('elapsed_time' in df.columns for df in optimization_data.values()):
        comparison_specs.insert(2, ('Elapsed Time (s)', 'elapsed_time', '{:.2f}'))

    for sheet_name, column_name, number_format in comparison_specs:
        create_comparison_sheet(wb, sheet_name, column_name, optimization_data, number_format)

    # Auto-size columns
    for sheet in wb.worksheets:
        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            sheet.column_dimensions[column_letter].width = max_length + 2

    wb.save(output_path)
    wb.close()
    print(f"Comparison Excel workbook saved to: {output_path}")


def create_comparison_sheet(wb, sheet_name, column_name, data, number_format):
    """
    Create a single comparison sheet in the workbook.

    Parameters
    ----------
    wb : openpyxl.Workbook
        Workbook to add sheet to
    sheet_name : str
        Name of the sheet
    column_name : str
        Column name from optimization_summary.csv to extract
    data : dict
        {case_name: optimization_summary_df, ...}
    number_format : str
        Format string for cell values (e.g., '{:.6e}', '{:.2f}')
    """
    ws = wb.create_sheet(sheet_name)

    # Header row with styling
    ws['A1'] = 'Iteration'
    ws['A1'].font = Font(bold=True)
    ws['A1'].fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')

    case_names = list(data.keys())
    for col_idx, case_name in enumerate(case_names, start=2):
        cell = ws.cell(1, col_idx, case_name)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')

    # Data rows
    max_iterations = max(len(df) for df in data.values())

    for row_idx in range(max_iterations):
        # Iteration number in column A
        ws.cell(row_idx + 2, 1, row_idx + 1)

        # Values for each case in subsequent columns
        for col_idx, (case_name, df) in enumerate(data.items(), start=2):
            if row_idx < len(df) and column_name in df.columns:
                value = df.iloc[row_idx][column_name]
                cell = ws.cell(row_idx + 2, col_idx)

                if isinstance(value, (int, float)):
                    cell.value = value
                    # Apply number format if not default
                    if number_format != '{}':
                        cell.number_format = number_format.replace('{:', '').replace('}', '')
                else:
                    cell.value = str(value)


def clean_column_names(df):
    """
    Extract simple variable names from CSV headers with format 'var_name, Description, (units)'.

    Parameters
    ----------
    df : pd.DataFrame
        DataFrame with full column names from results.csv

    Returns
    -------
    pd.DataFrame
        DataFrame with simplified column names (just the variable name)
    """
    new_columns = []
    for col in df.columns:
        var_name = col.split(',')[0].strip()
        new_columns.append(var_name)
    df.columns = new_columns
    return df


def create_results_comparison_xlsx(results_data, directories, output_path):
    """
    Create Excel workbook comparing results time series across cases.

    Creates multi-sheet workbook with one sheet per variable. Each sheet has
    time in column A and one column per case.

    Parameters
    ----------
    results_data : dict
        {case_name: results_df, ...}
    directories : list of Path
        List of result directories included in comparison
    output_path : Path or str
        Output Excel file path

    Notes
    -----
    Workbook structure:
    - Sheet 1: "Directories" - list of all compared directories
    - Subsequent sheets: One sheet per variable (25 total)
      - Column A: Time (years)
      - Columns B+: One column per case with variable values

    Variable sheets match the plots in comparison_plots.pdf:
    Economic: y, y_eff, K, Consumption, Savings, s, Y_gross, Y_net
    Climate: delta_T, E, Ecum
    Abatement/Damage: f, mu, Lambda, AbateCost, Omega, Climate_Damage
    Inequality/Utility: Gini, G_eff, U, discounted_utility
    Exogenous: A, L, sigma, theta1
    """
    if not results_data:
        print("No results data available - skipping results comparison workbook")
        return

    results_data_cleaned = {name: clean_column_names(df.copy()) for name, df in results_data.items()}

    variable_specs = [
        ('y', 'Per-Capita Consumption'),
        ('y_eff', 'Effective Per-Capita Income'),
        ('K', 'Capital Stock'),
        ('Consumption', 'Total Consumption'),
        ('Savings', 'Gross Investment'),
        ('s', 'Savings Rate'),
        ('Y_gross', 'Gross Production'),
        ('Y_net', 'Net Output After Damages & Abatement'),
        ('delta_T', 'Temperature Change from Pre-Industrial'),
        ('E', 'CO2 Emissions Rate'),
        ('Ecum', 'Cumulative CO2 Emissions'),
        ('f', 'Abatement Allocation Fraction'),
        ('mu', 'Emissions Abatement Fraction'),
        ('Lambda', 'Abatement Cost (% of Output)'),
        ('AbateCost', 'Total Abatement Cost'),
        ('Omega', 'Climate Damage (% of Output)'),
        ('Climate_Damage', 'Total Climate Damage'),
        ('Gini', 'Gini Index Before Redistribution'),
        ('G_eff', 'Effective Gini Index'),
        ('U', 'Mean Utility Per Capita'),
        ('discounted_utility', 'Discounted Utility Per Capita'),
        ('A', 'Total Factor Productivity'),
        ('L', 'Population'),
        ('sigma', 'Carbon Intensity of GDP'),
        ('theta1', 'Marginal Abatement Cost'),
    ]

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    create_directories_sheet(wb, directories)

    for var_name, sheet_title in variable_specs:
        if not any(var_name in df.columns for df in results_data_cleaned.values()):
            continue

        ws = wb.create_sheet(sheet_title[:31])

        ws['A1'] = 'Time (years)'
        ws['A1'].font = Font(bold=True)
        ws['A1'].fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')

        case_names = list(results_data_cleaned.keys())
        for col_idx, case_name in enumerate(case_names, start=2):
            cell = ws.cell(1, col_idx, case_name)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')

        representative_df = next(iter(results_data_cleaned.values()))
        time_values = representative_df['t'].values

        for row_idx, time_val in enumerate(time_values, start=2):
            ws.cell(row_idx, 1, float(time_val))

            for col_idx, (case_name, df) in enumerate(results_data_cleaned.items(), start=2):
                if var_name in df.columns:
                    value = df.iloc[row_idx - 2][var_name]
                    ws.cell(row_idx, col_idx, float(value))

    for sheet in wb.worksheets:
        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            sheet.column_dimensions[column_letter].width = max_length + 2

    wb.save(output_path)
    wb.close()
    print(f"Results comparison workbook saved to: {output_path}")
